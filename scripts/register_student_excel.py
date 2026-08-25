#!/usr/bin/env python3
"""
Python Script to Register a New Student and Append to Student_Improvement_Tracker.xlsx
"""

import sys, os, json
import openpyxl

EXCEL_PATH = '/working_dir/student-portal-repo/assets/data/Student_Improvement_Tracker.xlsx'
JSON_PATH = '/working_dir/student-portal-repo/assets/data/students.json'

def register_student(name, student_id, username, password, grade_class="06 - Science", teacher="Mrs. Sheshadi Sathsarani"):
    # 1. Update JSON
    if os.path.exists(JSON_PATH):
        with open(JSON_PATH, 'r') as f:
            students = json.load(f)
    else:
        students = []

    MONTHS = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    monthly_progress = {}
    for m in MONTHS:
        monthly_progress[m] = [
            {"week": "1 week", "master_guide_1": "Completed", "master_guide_2": "Completed", "past_paper": "Completed", "practical": "Good", "unit_test": 75},
            {"week": "2 week", "master_guide_1": "Completed", "master_guide_2": "Incomplete", "past_paper": "Incomplete", "practical": "Good", "unit_test": 70},
            {"week": "3 week", "master_guide_1": "Completed", "master_guide_2": "Completed", "past_paper": "Completed", "practical": "Excellent", "unit_test": 85},
            {"week": "4 week", "master_guide_1": "Completed", "master_guide_2": "Completed", "past_paper": "Completed", "practical": "Excellent", "unit_test": 90}
        ]

    new_st = {
        "tab_name": f"{name.split()[0]} {student_id}",
        "student_info": {
            "name": name,
            "student_id": student_id,
            "username": username.lower(),
            "password": password,
            "grade_class": grade_class,
            "homeroom_teacher": teacher,
            "avatar": f"https://api.dicebear.com/7.x/avataaars/svg?seed={name}",
            "qr_code_key": f"QR-{student_id}",
            "access_url": f"student.html?id={student_id}"
        },
        "weekly_progress": monthly_progress["January"],
        "monthly_progress": monthly_progress,
        "assessments": [
            {"term": "Term 1 Exam", "score": 78},
            {"term": "Term 2 Exam", "score": 82},
            {"term": "Final Exam", "score": 88}
        ],
        "summary": {
            "attendance": "95%",
            "average_unit_test": 80.0,
            "overall_status": "Active Progress"
        },
        "teacher_notes": "Registered via Excel script."
    }

    students.append(new_st)
    with open(JSON_PATH, 'w') as f:
        json.dump(students, f, indent=2)

    # 2. Append worksheet tab to Excel file if available
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        sheet_title = f"{name.split()[0]} ({student_id})"
        if sheet_title in wb.sheetnames:
            ws = wb[sheet_title]
        else:
            ws = wb.create_sheet(title=sheet_title)

        ws['A1'] = "Student Name:"
        ws['B1'] = name
        ws['A2'] = "Student ID:"
        ws['B2'] = student_id
        ws['A3'] = "Username:"
        ws['B3'] = username
        ws['A4'] = "Grade / Class:"
        ws['B4'] = grade_class

        headers = ["Weeks", "Master Guide 1", "Master Guide 2", "Past Paper", "Practical", "Unit Test"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=6, column=col_idx, value=h)

        row_num = 7
        for w in monthly_progress["January"]:
            ws.cell(row=row_num, column=1, value=w["week"])
            ws.cell(row=row_num, column=2, value=w["master_guide_1"])
            ws.cell(row=row_num, column=3, value=w["master_guide_2"])
            ws.cell(row=row_num, column=4, value=w["past_paper"])
            ws.cell(row=row_num, column=5, value=w["practical"])
            ws.cell(row=row_num, column=6, value=w["unit_test"])
            row_num += 1

        wb.save(EXCEL_PATH)
        print(f"Successfully appended student tab '{sheet_title}' to {EXCEL_PATH}")

if __name__ == '__main__':
    register_student("Samantha Perera", "ST-90001", "samantha", "samantha123")
